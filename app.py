
import streamlit as st

import re
import time
import requests
from datetime import datetime
import tempfile
from pathlib import Path
import os
from copy import copy
import json

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


# =========================
# Sanitización Excel (evita IllegalCharacterError)
# =========================
_ILLEGAL_XL_CHARS_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")

def excel_safe_value(v):
    if v is None:
        return None
    if isinstance(v, (int, float, bool)):
        return v
    if isinstance(v, (dict, list, tuple, set)):
        try:
            v = json.dumps(v, ensure_ascii=False, separators=(",", ":"))
        except Exception:
            v = str(v)
    if not isinstance(v, str):
        v = str(v)
    v = _ILLEGAL_XL_CHARS_RE.sub("", v)
    v = v.replace("\r\n", "\n").replace("\r", "\n")
    if len(v) > 32767:
        v = v[:32767]
    return v


# =========================
# CONFIG STREAMLIT
# =========================
st.set_page_config(page_title="DevengosCuentas2026", layout="wide")
st.title("App DevengosCuentas2026")
st.caption("BUILD: 9640a48 - 2026-02-04")

try:
    API_KEY = (st.secrets.get("API_KEY", "") or "").strip()
except Exception:
    API_KEY = ""

if not API_KEY:
    st.error("Falta API_KEY en Secrets (Streamlit Cloud → Manage app → Settings → Secrets).")
    st.stop()


# =========================
# (PROGRAMA 1)
# =========================
MAX_ITEMS = 20

API_COLUMNAS = [
    "Proveedor_Nombre",
    "Proveedor_Rut",
    "Codigo_OC",
    "N_Licitacion",
    "Descripcion_OC",
    "FechaCreacion_OC",
    "TotalNeto_OC",
    "NombreContacto_OC",
    "CantidadItems_OC",
]

for i in range(1, MAX_ITEMS + 1):
    API_COLUMNAS.extend([
        f"Cantidad_{i}",
        f"EspecificacionComprador_{i}",
        f"PrecioNeto_{i}",
    ])

COLUMNAS_ORDEN = ["Fecha", "Folio", "Título", "Monto"] + API_COLUMNAS


def copiar_estilo(origen, destino):
    destino.font = copy(origen.font)
    destino.border = copy(origen.border)
    destino.fill = copy(origen.fill)
    destino.number_format = origen.number_format
    destino.alignment = copy(origen.alignment)


def limpiar_folio(serie):
    return serie.astype(float).astype(int).astype(str)


def normalizar_monto(serie):
    return (
        serie.astype(str)
        .str.replace("(", "-", regex=False)
        .str.replace(")", "", regex=False)
        .str.strip()
        .astype(float)
    )


def leer_sigfe_nuevo(ruta_excel):
    df = pd.read_excel(ruta_excel, skiprows=5)
    df = df[df["Tipo Vista"] == "Tipo Flujo"]

    df = df.rename(columns={
        df.columns[1]: "Monto_Total_Cuenta",
        df.columns[-1]: "Monto"
    })

    df["Cuenta"] = df["Concepto"].astype(str).str.extract(r"(\d+)")
    df["Nombre_Cuenta"] = (
        df["Concepto"]
        .astype(str)
        .str.replace(r"\d+", "", regex=True)
        .str.replace("-", "", regex=False)
        .str.strip()
    )

    df["Fecha"] = pd.to_datetime(df["Fecha"]).dt.date

    df = df[[
        "Cuenta",
        "Nombre_Cuenta",
        "Fecha",
        "Folio",
        "Título",
        "Monto"
    ]].copy()

    df["Monto"] = normalizar_monto(df["Monto"])
    df["Folio"] = limpiar_folio(df["Folio"])

    for col in API_COLUMNAS:
        df[col] = ""

    return df


def leer_maestro(ruta_maestro):
    base_cols = ["Cuenta", "Fecha", "Folio", "Título", "Monto"] + API_COLUMNAS

    if not os.path.exists(ruta_maestro):
        return pd.DataFrame(columns=base_cols)

    xls = pd.ExcelFile(ruta_maestro)
    dfs = []

    for sheet in xls.sheet_names:
        temp = pd.read_excel(xls, sheet_name=sheet, skiprows=0)

        columnas_min = {"Fecha", "Folio", "Monto"}
        if not columnas_min.issubset(temp.columns):
            continue

        cols = list(temp.columns)
        if len(cols) >= 3:
            cols[2] = "Título"
            temp.columns = cols

        temp = temp.copy()
        temp["Cuenta"] = str(sheet)

        temp = temp[temp["Folio"].notna()]
        temp["Folio"] = limpiar_folio(temp["Folio"])

        temp["Fecha"] = pd.to_datetime(temp["Fecha"], errors="coerce").dt.date
        temp = temp[temp["Fecha"].notna()]

        for col in API_COLUMNAS:
            if col not in temp.columns:
                temp[col] = ""

        if "Título" not in temp.columns:
            temp["Título"] = ""

        dfs.append(temp)

    if not dfs:
        return pd.DataFrame(columns=base_cols)

    return pd.concat(dfs, ignore_index=True)


def obtener_devengos_nuevos(df_nuevo, df_maestro):
    if df_maestro.empty:
        return df_nuevo.copy()

    df_maestro = df_maestro.copy()
    df_nuevo = df_nuevo.copy()

    df_maestro["Cuenta"] = df_maestro["Cuenta"].astype(str)
    df_nuevo["Cuenta"] = df_nuevo["Cuenta"].astype(str)

    df_maestro["Folio"] = limpiar_folio(df_maestro["Folio"])
    df_nuevo["Folio"] = limpiar_folio(df_nuevo["Folio"])

    claves_maestro = set(
        df_maestro[["Cuenta", "Folio"]]
        .drop_duplicates()
        .itertuples(index=False, name=None)
    )

    claves_nuevo = df_nuevo[["Cuenta", "Folio"]].itertuples(index=False, name=None)
    mask = [clave not in claves_maestro for clave in claves_nuevo]

    return df_nuevo.loc[mask].copy()


def aplicar_formato_corporativo(wb, df_total):
    header_font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
    body_font = Font(name="Calibri", size=9)
    header_fill = PatternFill("solid", fgColor="1F4E78")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for ws in wb.worksheets:
        cuenta = ws.title

        nombre_cuenta = ""
        if "Nombre_Cuenta" in df_total.columns:
            aux = df_total[
                (df_total["Cuenta"] == cuenta) &
                (df_total["Nombre_Cuenta"].notna())
            ]["Nombre_Cuenta"]
            if not aux.empty:
                nombre_cuenta = aux.iloc[0]

        header_titulo = f"DEVENGOS 2026 - CUENTA {cuenta} - {nombre_cuenta}"

        for j, colname in enumerate(COLUMNAS_ORDEN, start=1):
            c = ws.cell(row=1, column=j)
            c.value = header_titulo if j == 3 else colname
            c.font = header_font
            c.fill = header_fill
            c.border = thin_border
            c.alignment = Alignment(horizontal="center", vertical="center")

        ws.freeze_panes = "A2"

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(COLUMNAS_ORDEN)):
            for cell in row:
                cell.font = body_font
                cell.border = thin_border

        for cell in ws["A"][1:]:
            cell.number_format = "DD-MM-YYYY"
        for cell in ws["D"][1:]:
            cell.number_format = "#,##0"

        for col_cells in ws.iter_cols(min_row=1, max_row=ws.max_row, max_col=len(COLUMNAS_ORDEN)):
            max_length = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 2, 70)


def ejecutar_programa_1(ruta_nuevo, ruta_maestro):
    df_nuevo = leer_sigfe_nuevo(ruta_nuevo)
    df_maestro = leer_maestro(ruta_maestro)

    archivo_nuevo = not os.path.exists(ruta_maestro)
    df_nuevos = obtener_devengos_nuevos(df_nuevo, df_maestro)

    if not os.path.exists(ruta_maestro):
        with pd.ExcelWriter(ruta_maestro, engine="openpyxl") as writer:
            for cuenta in df_nuevos["Cuenta"].unique():
                datos = df_nuevos[df_nuevos["Cuenta"] == cuenta].drop(columns=["Cuenta", "Nombre_Cuenta"])
                datos = datos.reindex(columns=COLUMNAS_ORDEN)
                datos.to_excel(writer, sheet_name=str(cuenta), index=False)

    wb = load_workbook(ruta_maestro)

    if not archivo_nuevo:
        for cuenta in df_nuevos["Cuenta"].unique():
            ws = wb[cuenta] if cuenta in wb.sheetnames else wb.create_sheet(cuenta)
            nuevos = df_nuevos[df_nuevos["Cuenta"] == cuenta]

            if ws.max_row < 1:
                for j, colname in enumerate(COLUMNAS_ORDEN, start=1):
                    ws.cell(row=1, column=j).value = colname

            fila_modelo = 2 if ws.max_row >= 2 else 1

            for _, row in nuevos.iterrows():
                fila_nueva = ws.max_row + 1
                valores = [row.get(col, "") for col in COLUMNAS_ORDEN]

                for col_idx, valor in enumerate(valores, start=1):
                    celda_nueva = ws.cell(row=fila_nueva, column=col_idx, value=valor)
                    celda_modelo = ws.cell(row=fila_modelo, column=col_idx)
                    copiar_estilo(celda_modelo, celda_nueva)

    df_total = df_nuevos.copy() if df_maestro.empty else pd.concat([df_maestro, df_nuevos], ignore_index=True)
    aplicar_formato_corporativo(wb, df_total)
    wb.save(ruta_maestro)


# =========================
# (PROGRAMA 2)
# =========================
TIMEOUT = 25
MAX_REINTENTOS = 4
PAUSA_ENTRE_LLAMADAS = 0.75

MAX_FALLOS_OC = 3
PREFIJOS_OC_OMITIR = ("621-",)
#(si mañana quieres omitir otras, agregas: ("621-", "622-", "900-"))

def normalizar_guiones(s: str) -> str:
    return (s.replace("\u2010", "-")
             .replace("\u2011", "-")
             .replace("\u2012", "-")
             .replace("\u2013", "-")
             .replace("\u2212", "-"))

def extraer_codigo_oc(texto):
    """
    Extrae OC incluso si viene con espacios: 1057547 - 588 - SE26
    o guiones raros (normalizados).
    """
    if texto is None:
        return None
    s = normalizar_guiones(str(texto))
    m = re.search(r"\b(\d{3,})\s*-\s*(\d{1,})\s*-\s*([A-Za-z]{1,6}\d{1,4})\b", s)
    if not m:
        return None
    return f"{m.group(1)}-{m.group(2)}-{m.group(3).upper()}"

def mapear_headers_fila1(ws):
    headers = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v is None:
            continue
        k = str(v).strip()
        if not k:
            continue
        headers[k] = col
    return headers

def encontrar_columna_texto_oc(headers):
    # 1) Caso ideal
    if "Título" in headers:
        return headers["Título"]

    # 2) Tu caso: el header de la col 3 tiene "DEVENGOS 2026 - CUENTA ..."
    for k, col in headers.items():
        if str(k).strip().upper().startswith("DEVENGOS 2026 - CUENTA"):
            return col

    # 3) Fallback: en tu layout siempre es la columna 3
    return 3

def obtener_datos_oc(session, codigo_oc):
    url = "https://api.mercadopublico.cl/servicios/v1/publico/ordenesdecompra.json"
    params = {"codigo": codigo_oc, "ticket": API_KEY}
    last_err = None

    for intento in range(1, MAX_REINTENTOS + 1):
        try:
            r = session.get(url, params=params, timeout=TIMEOUT)
            status = r.status_code

            try:
                resp = r.json()
            except Exception:
                last_err = f"respuesta no-JSON (status {status})"
                resp = None

            if status != 200:
                last_err = f"status {status} body={r.text[:200]}"
            elif not resp:
                last_err = f"respuesta vacía (status {status})"
            else:
                if resp.get("Listado"):
                    o = resp["Listado"][0]
                    datos = {}

                    datos["Codigo_OC"] = o.get("Codigo") or codigo_oc
                    datos["N_Licitacion"] = o.get("CodigoLicitacion") or ""

                    datos["Descripcion_OC"] = (o.get("Descripcion", "") or "").replace("\n", " ").replace("\r", " ")

                    fecha = o.get("Fechas", {}).get("FechaCreacion")
                    datos["FechaCreacion_OC"] = (
                        datetime.fromisoformat(fecha.split(".")[0]).strftime("%d/%m/%Y")
                        if fecha else ""
                    )

                    datos["TotalNeto_OC"] = o.get("TotalNeto")
                    datos["NombreContacto_OC"] = (o.get("Comprador", {}) or {}).get("NombreContacto") or ""
                    datos["CantidadItems_OC"] = (o.get("Items", {}) or {}).get("Cantidad") or 0

                    prov = o.get("Proveedor", {}) or {}
                    datos["Proveedor_Nombre"] = prov.get("Nombre", "") or ""
                    datos["Proveedor_Rut"] = prov.get("RutSucursal") or prov.get("Rut") or ""

                    items = (o.get("Items", {}) or {}).get("Listado", []) or []
                    for i, item in enumerate(items[:MAX_ITEMS], start=1):
                        datos[f"Cantidad_{i}"] = item.get("Cantidad")
                        datos[f"EspecificacionComprador_{i}"] = (item.get("EspecificacionComprador", "") or "").replace("\n", " ").replace("\r", " ")
                        datos[f"PrecioNeto_{i}"] = item.get("PrecioNeto")

                    return datos, None

                msg = resp.get("Mensaje") or resp.get("message") or resp.get("Error") or ""
                last_err = f"sin Listado (status {status}) {msg}".strip()

        except requests.RequestException as e:
            last_err = f"request error: {type(e).__name__}"

        time.sleep(0.8 * intento)

    return None, last_err or "sin datos"

def escribir_hoja_errores(wb, errores):
    nombre = "ERRORES_API"
    headers_err = ["Hoja", "Fila", "Codigo_OC", "Motivo", "Detalle", "Faltan", "Titulo", "Timestamp"]

    if nombre in wb.sheetnames:
        ws_err = wb[nombre]
        # Si está vacía o sin headers, crear headers
        if ws_err.max_row < 1 or (ws_err.cell(row=1, column=1).value != "Hoja"):
            ws_err.delete_rows(1, ws_err.max_row)
            ws_err.append(headers_err)
            ws_err.freeze_panes = "A2"
    else:
        ws_err = wb.create_sheet(nombre)
        ws_err.append(headers_err)
        ws_err.freeze_panes = "A2"

    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for e in errores:
        ws_err.append([
            excel_safe_value(e.get("Hoja", "")),
            excel_safe_value(e.get("Fila", "")),
            excel_safe_value(e.get("Codigo_OC", "")),
            excel_safe_value(e.get("Motivo", "")),
            excel_safe_value(e.get("Detalle", "")),
            excel_safe_value(e.get("Faltan", "")),
            excel_safe_value(e.get("Titulo", "")),
            ts,
        ])

def ejecutar_programa_2(ruta_maestro, hojas_permitidas=None):
    wb = load_workbook(ruta_maestro)
    errores = []

    HOJAS_OMITIDAS = {"220400400101", "220400400102"}

    cache_ok = {}
    cache_fail = {}  # codigo_oc -> cantidad_fallos

    session = requests.Session()

    CAMPOS_CLAVE_API = ["Proveedor_Nombre", "Proveedor_Rut", "N_Licitacion", "Descripcion_OC", "FechaCreacion_OC"]

    for ws in wb.worksheets:
        if ws.title in HOJAS_OMITIDAS:
            continue
        if hojas_permitidas is not None and ws.title not in hojas_permitidas:
            continue

        headers = mapear_headers_fila1(ws)

        # ✅ Encontrar la columna con el texto donde viene la OC (aunque el header sea DEVENGOS...)
        col_titulo = encontrar_columna_texto_oc(headers)

        if "Codigo_OC" not in headers:
            continue
        col_codigo = headers["Codigo_OC"]

        fila = 2
        while fila <= ws.max_row:
            titulo = ws.cell(row=fila, column=col_titulo).value or ""

            codigo_oc = extraer_codigo_oc(titulo)
            if not codigo_oc:
                errores.append({
                    "Hoja": ws.title,
                    "Fila": fila,
                    "Motivo": "NO_SE_ENCONTRO_OC_EN_TEXTO",
                    "Titulo": str(titulo)[:500],
                })
                fila += 1
                continue

            # ✅ Si la OC está en lista de omitidas, NO consultar API (por decisión de negocio)
            if codigo_oc.startswith(PREFIJOS_OC_OMITIR):
                # (Opcional) dejar registro en ERRORES_API para auditoría
                errores.append({
                    "Hoja": ws.title,
                    "Fila": fila,
                    "Codigo_OC": codigo_oc,
                    "Motivo": "OC_OMITIDA_POR_REGLA",
                    "Detalle": f"Se omitió consulta API para prefijos: {PREFIJOS_OC_OMITIR}",
                    "Titulo": str(titulo)[:500],
                })

                # Igual dejar escrito el Codigo_OC si estaba vacío
                celda_cod = ws.cell(row=fila, column=col_codigo)
                if celda_cod.value in (None, ""):
                    celda_cod.value = excel_safe_value(codigo_oc)

                fila += 1
                continue





            # ✅ Si hay OC en el texto, siempre dejarla escrita en Codigo_OC
            celda_cod = ws.cell(row=fila, column=col_codigo)
            if celda_cod.value in (None, ""):
                celda_cod.value = excel_safe_value(codigo_oc)

            # Completar solo si faltan campos clave
            faltan = []
            for c in CAMPOS_CLAVE_API:
                if c in headers:
                    v = ws.cell(row=fila, column=headers[c]).value
                    if v in (None, ""):
                        faltan.append(c)

            if not faltan:
                fila += 1
                continue

            # Cache + control de fallos
            if codigo_oc in cache_ok:
                datos = cache_ok[codigo_oc]
                err = None
            else:
                if cache_fail.get(codigo_oc, 0) >= MAX_FALLOS_OC:
                    errores.append({
                        "Hoja": ws.title,
                        "Fila": fila,
                        "Codigo_OC": codigo_oc,
                        "Motivo": "API_SKIPPED_POR_MUCHOS_FALLOS",
                        "Detalle": f"fallos={cache_fail.get(codigo_oc, 0)}",
                        "Faltan": ", ".join(faltan),
                        "Titulo": str(titulo)[:500],
                    })
                    fila += 1
                    continue

                time.sleep(PAUSA_ENTRE_LLAMADAS)
                datos, err = obtener_datos_oc(session, codigo_oc)

                if datos:
                    cache_ok[codigo_oc] = datos
                else:
                    cache_fail[codigo_oc] = cache_fail.get(codigo_oc, 0) + 1
                    errores.append({
                        "Hoja": ws.title,
                        "Fila": fila,
                        "Codigo_OC": codigo_oc,
                        "Motivo": "API_FAIL",
                        "Detalle": (err or "")[:500],
                        "Faltan": ", ".join(faltan),
                        "Titulo": str(titulo)[:500],
                    })
                    fila += 1
                    continue

            # Log si API OK pero aún faltan datos clave
            faltan_post = []
            for c in faltan:
                v = (datos or {}).get(c)
                if v in (None, "", "No disponible"):
                    faltan_post.append(c)
            if faltan_post:
                errores.append({
                    "Hoja": ws.title,
                    "Fila": fila,
                    "Codigo_OC": codigo_oc,
                    "Motivo": "API_OK_PERO_FALTAN_DATOS",
                    "Faltan": ", ".join(faltan_post),
                    "Titulo": str(titulo)[:500],
                })

            # Escribir columnas existentes
            for campo, valor in (datos or {}).items():
                if campo in headers:
                    ws.cell(row=fila, column=headers[campo]).value = excel_safe_value(valor)

            fila += 1

    escribir_hoja_errores(wb, errores)
    wb.save(ruta_maestro)


# =========================
# UI + DESCARGA
# =========================
sigfe_file = st.file_uploader("1) Subir SA_MayorPresupuestario.xls", type=["xls", "xlsx"])
maestro_file = st.file_uploader("2) (Opcional) Subir DevengosCuentas2026.xlsx existente", type=["xlsx"])

# --- Session state ---
if "excel_bytes_p1" not in st.session_state:
    st.session_state.excel_bytes_p1 = None
if "hojas_generadas" not in st.session_state:
    st.session_state.hojas_generadas = []
if "excel_bytes_final" not in st.session_state:
    st.session_state.excel_bytes_final = None

# ✅ Flag anti “se ejecuta solo”
if "run_p2" not in st.session_state:
    st.session_state.run_p2 = False

# ✅ Guardar selección en session_state (para que no se pierda)
if "sel_hojas_api" not in st.session_state:
    st.session_state.sel_hojas_api = []

EXCLUIDAS = {"220400400101", "220400400102"}

colA, colB = st.columns(2)

with colA:
    if st.button("1️⃣ Generar Excel (Programa 1 XD232)"):
        if sigfe_file is None:
            st.error("Falta SA_MayorPresupuestario.xls")
            st.stop()

        with st.spinner("Generando Devengos (Programa 1)..."):
            with tempfile.TemporaryDirectory() as tmpdir:
                tmp = Path(tmpdir)

                ruta_sigfe = tmp / "SA_MayorPresupuestario.xls"
                ruta_sigfe.write_bytes(sigfe_file.getbuffer())

                ruta_maestro = tmp / "DevengosCuentas2026.xlsx"
                if maestro_file is not None:
                    ruta_maestro.write_bytes(maestro_file.getbuffer())

                ejecutar_programa_1(str(ruta_sigfe), str(ruta_maestro))

                st.session_state.excel_bytes_p1 = ruta_maestro.read_bytes()

                wb_tmp = load_workbook(str(ruta_maestro), read_only=True)
                hojas = [h for h in wb_tmp.sheetnames if h not in EXCLUIDAS]
                st.session_state.hojas_generadas = hojas

                st.session_state.excel_bytes_final = None
                st.session_state.sel_hojas_api = []  # reset selección

                st.success(f"✅ Programa 1 listo. Pestañas generadas: {len(hojas)}")

st.markdown("### 2️⃣ Selecciona las cuentas/pestañas a completar xDXD1")

if not st.session_state.get("excel_bytes_p1"):
    st.info("Primero ejecuta **Programa 1** para generar el Excel.")
else:
    # =========================
    # Selección de cuentas (Programa 2) - FIX session_state
    # =========================
    hojas = st.session_state.get("hojas_generadas", [])

    # Callbacks (única forma válida de modificar el state del multiselect)
    def cb_sel_todas():
        st.session_state["sel_hojas_api"] = hojas

    def cb_limpiar():
        st.session_state["sel_hojas_api"] = []

    # Multiselect (NO tocar session_state directamente después)
    seleccion = st.multiselect(
        "Escribe para buscar y selecciona las cuentas a completar (Programa 2)",
        options=hojas,
        default=st.session_state.get("sel_hojas_api", []),
        key="sel_hojas_api"
    )

    c1, c2 = st.columns(2)
    with c1:
        st.button(
            "✅ Seleccionar todas",
            key="btn_sel_todas",
            on_click=cb_sel_todas
        )
    with c2:
        st.button(
            "🧹 Limpiar",
            key="btn_limpiar",
            on_click=cb_limpiar
        )

    if st.button("2️⃣ Completar API (Programa 2)", key="btn_run_p2"):
        if not st.session_state.sel_hojas_api:
            st.warning("No seleccionaste ninguna cuenta.")
        else:
            st.session_state.run_p2 = True
            st.rerun()

    # ✅ Ejecutar SOLO si el flag está encendido (y apagarlo altiro)
    if st.session_state.run_p2:
        st.session_state.run_p2 = False

        seleccion_set = set(st.session_state.sel_hojas_api)

        with st.spinner("Consultando API y completando Excel..."):
            with tempfile.TemporaryDirectory() as tmpdir:
                tmp = Path(tmpdir)
                ruta_maestro = tmp / "DevengosCuentas2026.xlsx"

                # ✅ Base correcta:
                # - si usuario subió Excel, usarlo
                # - si no, usar el generado por Programa 1
                if maestro_file is not None:
                    ruta_maestro.write_bytes(maestro_file.getbuffer())
                    base_usada = "archivo subido por el usuario"
                else:
                    ruta_maestro.write_bytes(st.session_state.excel_bytes_p1)
                    base_usada = "archivo generado por Programa 1"

                ejecutar_programa_2(str(ruta_maestro), hojas_permitidas=seleccion_set)

                st.session_state.excel_bytes_final = ruta_maestro.read_bytes()

        st.success(f"✅ Programa 2 listo ({base_usada}). Cuentas procesadas: {len(seleccion_set)}")

st.markdown("### 3️⃣ Descargar")

excel_actual = st.session_state.get("excel_bytes_final") or st.session_state.get("excel_bytes_p1")

if excel_actual:
    st.download_button(
        label="⬇️ Descargar DevengosCuentas2026.xlsx",
        data=excel_actual,
        file_name="DevengosCuentas2026.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="download_devengos"
    )
else:
    st.info("Aún no hay archivo para descargar. Ejecuta Programa 1.")


